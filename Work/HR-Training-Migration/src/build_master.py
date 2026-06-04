"""Build the reviewable master sheet.

Pipeline:  extract -> normalise -> compute expiry -> flag -> write Excel + CSV.

Usage:
    python src/build_master.py --input sample_data --output output/master.xlsx
    python src/build_master.py --input sample_data --today 2026-06-04

Output:
    master.xlsx with two sheets:
      - "Training Records": every record, colour-coded by status
      - "Review Queue":     rows needing a human (unmatched course / bad date /
                            low-confidence fuzzy match)
    master.csv: the same Training Records data, for the browser-entry step.
"""
from __future__ import annotations

import argparse
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import expiry as exp
from catalogue import load_catalogue
from extract import read_folder
from normalise import course_for, normalise_course, normalise_name, parse_completion_date

# status -> (fill colour, needs_review)
STATUS_FILL = {
    exp.OVERDUE: "FFC7CE",        # red
    exp.EXPIRING_SOON: "FFEB9C",  # amber
    exp.VALID: "C6EFCE",          # green
    exp.UNKNOWN: "D9D9D9",        # grey
}

OUTPUT_COLUMNS = [
    "staff_name", "course_canonical", "completion_date", "interval_months",
    "expiry_date", "days_to_expiry", "status", "match_confidence",
    "needs_review", "review_reason", "raw_course", "source",
]


def build_rows(df: pd.DataFrame, cat, today: date) -> pd.DataFrame:
    records = []
    for row in df.itertuples(index=False):
        name = normalise_name(getattr(row, "staff_name", None))
        canonical, confidence, raw_course = normalise_course(getattr(row, "course", None), cat)
        completion, date_note = parse_completion_date(getattr(row, "completion_date", None))

        reasons = []
        if not name:
            reasons.append("missing name")
        if canonical is None:
            reasons.append(f"unmatched course (best {confidence}%)")
        if completion is None:
            reasons.append(f"bad date ({date_note})")

        course = course_for(canonical, cat) if canonical else None
        interval = course.interval_months if course else cat.default_interval_months
        expiry_date = exp.compute_expiry(completion, interval)
        status = exp.status_for(expiry_date, today, cat.expiring_soon_days)
        days = exp.days_until_expiry(expiry_date, today)

        records.append({
            "staff_name": name,
            "course_canonical": canonical or "",
            "completion_date": completion.isoformat() if completion else "",
            "interval_months": interval,
            "expiry_date": expiry_date.isoformat() if expiry_date else "",
            "days_to_expiry": days if days is not None else "",
            "status": status,
            "match_confidence": confidence,
            "needs_review": bool(reasons),
            "review_reason": "; ".join(reasons),
            "raw_course": raw_course,
            "source": getattr(row, "source", ""),
        })
    out = pd.DataFrame.from_records(records, columns=OUTPUT_COLUMNS)
    # Sort: overdue first, then soonest expiry
    status_rank = {exp.OVERDUE: 0, exp.EXPIRING_SOON: 1, exp.VALID: 2, exp.UNKNOWN: 3}
    out["_rank"] = out["status"].map(status_rank).fillna(9)
    out = out.sort_values(["_rank", "expiry_date"]).drop(columns="_rank").reset_index(drop=True)
    return out


def write_excel(df: pd.DataFrame, path: Path) -> None:
    review = df[df["needs_review"]].copy()
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Training Records", index=False)
        review.to_excel(writer, sheet_name="Review Queue", index=False)
        _style_sheet(writer.sheets["Training Records"], df)
        _style_sheet(writer.sheets["Review Queue"], review)


def _style_sheet(ws, df: pd.DataFrame) -> None:
    # Bold header
    for cell in ws[1]:
        cell.font = Font(bold=True)
    if df.empty:
        return
    status_col = OUTPUT_COLUMNS.index("status") + 1
    for excel_row, status in enumerate(df["status"], start=2):
        colour = STATUS_FILL.get(status)
        if colour:
            ws.cell(row=excel_row, column=status_col).fill = PatternFill(
                start_color=colour, end_color=colour, fill_type="solid"
            )
    # Reasonable column widths
    for i, col in enumerate(df.columns, start=1):
        width = max(len(str(col)), *(len(str(v)) for v in df[col])) if len(df) else len(str(col))
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 45)


def main() -> None:
    ap = argparse.ArgumentParser(description="Build the HR training master sheet.")
    ap.add_argument("--input", required=True, help="Folder of source spreadsheets/CSVs")
    ap.add_argument("--output", default="output/master.xlsx", help="Output .xlsx path")
    ap.add_argument("--config", default="config/courses.yaml", help="Course catalogue YAML")
    ap.add_argument("--today", default=None, help="Override 'today' (YYYY-MM-DD) for testing")
    args = ap.parse_args()

    today = datetime.strptime(args.today, "%Y-%m-%d").date() if args.today else date.today()
    cat = load_catalogue(args.config)
    raw = read_folder(args.input)
    out = build_rows(raw, cat, today)

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    write_excel(out, out_path)
    out.to_csv(out_path.with_suffix(".csv"), index=False)

    n = len(out)
    review = int(out["needs_review"].sum())
    overdue = int((out["status"] == exp.OVERDUE).sum())
    soon = int((out["status"] == exp.EXPIRING_SOON).sum())
    print(f"Processed {n} records  ->  {out_path}")
    print(f"  Overdue: {overdue}  |  Expiring soon: {soon}  |  Need review: {review}")


if __name__ == "__main__":
    main()
