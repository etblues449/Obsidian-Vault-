"""Income forecast workbook generator.

Reads Feb 2026 actuals (hardcoded from the user's paste) and writes a workbook
with one sheet per month from Feb 2026 to Dec 2026.

  * Feb 26  -- pre-filled with the actual values supplied.
  * Mar 26 -> Dec 26 -- forecasted from each payer's weekday + frequency.
  * "Other HB and SC" and "Sandwell - HB" pay GBP 4000 every day.
  * Each sheet has weekday headers, dated columns, a Total column,
    and a TOTAL row with SUM formulas.

Run:  python3 scripts/forecast_tool.py
Output:  Income_forecast_2026.xlsx (in repo root)
"""

from __future__ import annotations

import calendar
import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


YEAR = 2026
MONTHS = list(range(2, 13))  # February through December

# Payer config: (name, frequency, weekday, amount, week_occurrences)
#   frequency       -- "weekly" | "fortnightly" | "monthly" | "daily"
#   weekday         -- 0=Mon .. 6=Sun, or None when no Feb data / daily payer
#   amount          -- GBP per individual payment (for daily, GBP per day)
#   week_occurrences -- for fortnightly/monthly, which Nth-of-weekday slots
PAYERS: list[tuple[str, str, int | None, float, list[int] | None]] = [
    ("Birmingham City Council - Supported Living", "weekly",      3, 59419.08,  None),
    ("Haringey",                                   "weekly",      None, 0,      None),
    ("Sandwell MBC - Respite",                     "fortnightly", 0, 41250.75,  [1, 3]),
    ("Sandwell MBC - Daycare",                     "fortnightly", 4, 6585.35,   [1, 3]),
    ("Wolverhampton - Respite",                    "monthly",     None, 0,      None),
    ("Wolverhampton - Daycare",                    "weekly",      3, 5606.30,   None),
    ("Walsall - Respite",                          "weekly",      1, 11460.80,  None),
    ("Walsall - Daycare",                          "weekly",      None, 0,      None),
    ("Ideal for All Ltd - Direct Payments",        "weekly",      4, 4131.62,   None),
    ("Kirklees - Supported Living",                "weekly",      1, 5458.48,   None),
    ("LB Hillingdon - Supported Living",           "weekly",      2, 5203.80,   None),
    ("NHS Birmingham & Solihull ICB",              "weekly",      3, 7814.17,   None),
    ("NHS Black Country ICB",                      "weekly",      3, 5499.45,   None),
    ("Sandwell - HB",                              "daily",       None, 4000,   None),
    ("Sandwell MBC - Supported Living",            "fortnightly", 0, 132834.80, [1, 4]),
    ("Solihull MBC - Supported Living",            "weekly",      6, 33031.60,  None),
    ("Walsall - Supported Living",                 "weekly",      1, 319655.76, None),
    ("Wolverhampton - Supported Living",           "weekly",      3, 132264.95, None),
    ("People Plus - Direct Payments",              "weekly",      3, 21527.20,  None),
    ("Middlesbrough - Supported Living",           "weekly",      3, 6115.40,   None),
    ("Wolverhampton CC - HB",                      "daily",       None, 4000,   None),
    ("Chasing",                                    "monthly",     4, 5266.40,   [1]),
    ("Other HB and SC",                            "daily",       None, 4000,   None),
]


# --- Feb 2026 actuals from the user's paste ------------------------------
FEB_ACTUALS: dict[tuple[str, int], float] = {
    ("Birmingham City Council - Supported Living", 5):  59419.08,
    ("Birmingham City Council - Supported Living", 25): 59419.00,
    ("Birmingham City Council - Supported Living", 27): 12166.56,
    ("Sandwell MBC - Respite", 2):  41250.75,
    ("Sandwell MBC - Daycare", 6):  6840.00,
    ("Sandwell MBC - Daycare", 18): 6330.69,
    ("Wolverhampton - Daycare", 5):  5606.30,
    ("Wolverhampton - Daycare", 12): 5606.30,
    ("Wolverhampton - Daycare", 19): 5606.30,
    ("Wolverhampton - Daycare", 26): 5606.30,
    ("Walsall - Respite", 10): 11460.80,
    ("Ideal for All Ltd - Direct Payments", 6): 4131.62,
    ("Kirklees - Supported Living", 24): 5458.48,
    ("LB Hillingdon - Supported Living", 18): 5203.80,
    ("NHS Birmingham & Solihull ICB", 12): 7814.17,
    ("NHS Black Country ICB", 12): 2996.16,
    ("NHS Black Country ICB", 18): 5533.04,
    ("NHS Black Country ICB", 26): 7969.15,
    ("Sandwell MBC - Supported Living", 2):  102346.91,
    ("Sandwell MBC - Supported Living", 26): 163322.69,
    ("Solihull MBC - Supported Living", 22): 33031.60,
    ("Walsall - Supported Living", 10): 319655.76,
    ("Wolverhampton - Supported Living", 5):  132264.95,
    ("Wolverhampton - Supported Living", 12): 132264.95,
    ("Wolverhampton - Supported Living", 19): 132264.95,
    ("Wolverhampton - Supported Living", 26): 132264.95,
    ("People Plus - Direct Payments", 5): 21527.20,
    ("Middlesbrough - Supported Living", 19): 6115.40,
    ("Chasing", 6): 5266.40,
}

OTHER_HB_FEB: dict[int, float] = {
    1: 1359.30,  2: 11194.12, 3: 3746.52,  4: 4707.86,  5: 3584.14,  6: 5178.66,  7: 0.00,
    8: 200.00,   9: 7878.87, 10: 4192.21, 11: 6284.53, 12: 631.08,  13: 7331.58, 14: 25.00,
    15: 0.00,   16: 8146.57, 17: 3889.46, 18: 2853.92, 19: 2853.92, 20: 6704.54, 21: 1475.00,
    22: 200.00, 23: 21454.19, 24: 1931.08, 25: 2158.19, 26: 2427.26, 27: 8918.60, 28: 1162.84,
}


# --- Helpers --------------------------------------------------------------
def days_in_month(year: int, month: int) -> int:
    return calendar.monthrange(year, month)[1]


def occurrences_of_weekday(year: int, month: int, weekday: int) -> list[int]:
    n = days_in_month(year, month)
    return [d for d in range(1, n + 1) if datetime.date(year, month, d).weekday() == weekday]


def expected_payment_days(year: int, month: int, freq: str,
                          weekday: int | None, week_occ: list[int] | None) -> list[int]:
    if freq == "daily":
        return list(range(1, days_in_month(year, month) + 1))
    if weekday is None:
        return []
    occ = occurrences_of_weekday(year, month, weekday)
    if freq == "weekly":
        return occ
    if freq == "fortnightly":
        return [occ[i - 1] for i in (week_occ or [1, 3]) if 0 <= i - 1 < len(occ)]
    if freq == "monthly":
        idx = (week_occ or [1])[0] - 1
        return [occ[idx]] if 0 <= idx < len(occ) else []
    return []


def col_for_day(day: int) -> str:
    return get_column_letter(day + 1)


def sheet_name_for(month: int) -> str:
    if month == 2:
        return "Feb 26"
    return f"{calendar.month_name[month]} 26"


# --- Workbook builder -----------------------------------------------------
def build_workbook(output: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    money_fmt   = '_-£* #,##0.00_-;-£* #,##0.00_-;_-£* "-"??_-;_-@_-'
    header_font = Font(bold=True, size=11)
    bold        = Font(bold=True)
    grand_font  = Font(bold=True, size=12)
    weekday_fill = PatternFill("solid", fgColor="DDEBF7")
    date_fill    = PatternFill("solid", fgColor="F2F2F2")
    total_fill   = PatternFill("solid", fgColor="FFF2CC")
    weekday_names = ["Monday", "Tuesday", "Wednesday", "Thursday",
                     "Friday", "Saturday", "Sunday"]

    for month in MONTHS:
        n_days   = days_in_month(YEAR, month)
        ws       = wb.create_sheet(sheet_name_for(month))
        total_col = n_days + 2

        # Row 1: weekday headers
        for d in range(1, n_days + 1):
            dt = datetime.date(YEAR, month, d)
            c = ws.cell(row=1, column=d + 1, value=weekday_names[dt.weekday()])
            c.font = header_font
            c.fill = weekday_fill
            c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=1, column=total_col, value="Total")
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

        # Row 2: dates
        for d in range(1, n_days + 1):
            dt = datetime.date(YEAR, month, d)
            c = ws.cell(row=2, column=d + 1, value=dt)
            c.number_format = "dd mmmm yyyy"
            c.fill = date_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center")

        # Payer rows
        for i, (name, freq, weekday, amount, week_occ) in enumerate(PAYERS):
            r = i + 3
            ws.cell(row=r, column=1, value=name).font = bold

            if month == 2:
                if name == "Other HB and SC":
                    for d, val in OTHER_HB_FEB.items():
                        c = ws.cell(row=r, column=d + 1, value=val)
                        c.number_format = money_fmt
                else:
                    for (pname, day), val in FEB_ACTUALS.items():
                        if pname == name:
                            c = ws.cell(row=r, column=day + 1, value=val)
                            c.number_format = money_fmt
            else:
                # Forecast: place amount on every expected payment day
                if amount > 0:
                    for d in expected_payment_days(YEAR, month, freq, weekday, week_occ):
                        c = ws.cell(row=r, column=d + 1, value=amount)
                        c.number_format = money_fmt

            # Row total formula
            first, last = col_for_day(1), col_for_day(n_days)
            tcell = ws.cell(row=r, column=total_col,
                            value=f"=SUM({first}{r}:{last}{r})")
            tcell.number_format = money_fmt
            tcell.font = bold

        # TOTAL row
        total_row = len(PAYERS) + 3
        tlabel = ws.cell(row=total_row, column=1, value="TOTAL")
        tlabel.font = grand_font
        tlabel.fill = total_fill
        for d in range(1, n_days + 1):
            letter = col_for_day(d)
            c = ws.cell(row=total_row, column=d + 1,
                        value=f"=SUM({letter}3:{letter}{total_row - 1})")
            c.number_format = money_fmt
            c.font = bold
            c.fill = total_fill
        gcol = get_column_letter(total_col)
        gc = ws.cell(row=total_row, column=total_col,
                     value=f"=SUM({gcol}3:{gcol}{total_row - 1})")
        gc.number_format = money_fmt
        gc.font = grand_font
        gc.fill = total_fill

        # Layout
        ws.column_dimensions["A"].width = 42
        for d in range(1, n_days + 1):
            ws.column_dimensions[col_for_day(d)].width = 13
        ws.column_dimensions[gcol].width = 16
        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 18
        ws.freeze_panes = "B3"

    wb.save(output)
    print(f"Wrote {output}")


if __name__ == "__main__":
    out = Path(__file__).resolve().parent.parent / "Income_forecast_2026.xlsx"
    build_workbook(out)
